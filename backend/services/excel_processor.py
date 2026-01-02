# -*- coding: utf-8 -*-
"""
Servicio Procesador de Excel - Extracción de servicios de ANEXO 1
Version 15.1 con validaciones mejoradas
"""
import pandas as pd
import re
import os
import zipfile
import threading
from typing import List, Dict, Tuple, Optional, Any, Set
from dataclasses import dataclass, field
from datetime import datetime
from difflib import SequenceMatcher

from config import PROCESSING_CONFIG

# ══════════════════════════════════════════════════════════════════════════════
# CONSTANTES Y LISTAS DE VALIDACIÓN v15.1
# ══════════════════════════════════════════════════════════════════════════════

CIUDADES_COLOMBIA_COMPLETA: Set[str] = {
    # Capitales
    'BOGOTÁ', 'BOGOTA', 'MEDELLÍN', 'MEDELLIN', 'CALI', 'BARRANQUILLA',
    'CARTAGENA', 'BUCARAMANGA', 'CÚCUTA', 'CUCUTA', 'PEREIRA', 'IBAGUÉ',
    'IBAGUE', 'SANTA MARTA', 'MANIZALES', 'VILLAVICENCIO', 'PASTO',
    'MONTERÍA', 'MONTERIA', 'NEIVA', 'ARMENIA', 'SINCELEJO', 'POPAYÁN',
    'POPAYAN', 'VALLEDUPAR', 'TUNJA', 'FLORENCIA', 'QUIBDÓ', 'QUIBDO',
    'RIOHACHA', 'YOPAL', 'MOCOA', 'LETICIA', 'INÍRIDA', 'INIRIDA',
    'MITÚ', 'MITU', 'PUERTO CARREÑO', 'SAN JOSÉ DEL GUAVIARE', 'ARAUCA',
    # Ciudades intermedias usadas en traslados
    'BAHIA SOLANO', 'BARRANCABERMEJA', 'BUENAVENTURA', 'PALMIRA',
    'CARTAGO', 'TULUA', 'TULUÁ', 'BUGA', 'SOGAMOSO', 'DUITAMA', 'GIRARDOT',
    'FUSAGASUGA', 'FUSAGASUGÁ', 'FACATATIVA', 'FACATATIVÁ', 'ZIPAQUIRA',
    'ZIPAQUIRÁ', 'CHIA', 'CHÍA', 'SOACHA', 'RIONEGRO', 'ENVIGADO',
    'ITAGUI', 'ITAGÜÍ', 'BELLO', 'TUMACO', 'IPIALES', 'GRANADA', 'ACACIAS',
    'ACACÍAS', 'PUERTO LOPEZ', 'PUERTO LÓPEZ', 'AGUACHICA', 'OCAÑA',
    'APARTADO', 'APARTADÓ', 'TURBO', 'CAUCASIA', 'MAGANGUE', 'MAGANGUÉ',
    'LORICA', 'CERETE', 'CERETÉ', 'ESPINAL', 'MELGAR', 'FLANDES', 'HONDA',
    'MARIQUITA', 'LA DORADA', 'PUERTO BERRIO', 'PUERTO BERRÍO',
    'PUERTO BOYACA', 'PUERTO BOYACÁ', 'CIENAGA', 'CIÉNAGA', 'FUNDACION',
    'FUNDACIÓN', 'ARACATACA', 'EL BANCO', 'PLATO', 'COROZAL', 'SAMPUES',
    'SAMPUÉS', 'SAN MARCOS', 'ZARZAL', 'JAMUNDI', 'JAMUNDÍ', 'YUMBO',
    'CANDELARIA', 'PRADERA', 'FLORIDA', 'CERRITO', 'GUACARI', 'GUACARÍ',
    'GINEBRA', 'ROLDANILLO', 'LA UNION', 'LA UNIÓN', 'SEVILLA',
    'CAICEDONIA', 'ARGELIA', 'DARIEN', 'DARIÉN', 'RESTREPO', 'DAGUA',
    'LA CUMBRE', 'CLO', 'BOG', 'MDE',  # Códigos de aeropuerto
}

DEPARTAMENTOS_COLOMBIA: Set[str] = {
    'BOGOTÁ D.C', 'BOGOTA D.C', 'BOGOTÁ D.C.', 'BOGOTA D.C.',
    'ANTIOQUIA', 'ATLÁNTICO', 'ATLANTICO', 'BOLÍVAR', 'BOLIVAR',
    'BOYACÁ', 'BOYACA', 'CALDAS', 'CAQUETÁ', 'CAQUETA', 'CASANARE',
    'CAUCA', 'CESAR', 'CHOCÓ', 'CHOCO', 'CÓRDOBA', 'CORDOBA',
    'CUNDINAMARCA', 'GUAINÍA', 'GUAINIA', 'GUAVIARE', 'HUILA',
    'LA GUAJIRA', 'MAGDALENA', 'META', 'NARIÑO', 'NARINO',
    'NORTE DE SANTANDER', 'PUTUMAYO', 'QUINDÍO', 'QUINDIO',
    'RISARALDA', 'SAN ANDRÉS', 'SAN ANDRES', 'SANTANDER', 'SUCRE',
    'TOLIMA', 'VALLE', 'VALLE DEL CAUCA', 'VAUPÉS', 'VAUPES',
    'VICHADA', 'AMAZONAS', 'ARAUCA'
}

MUNICIPIOS_COLOMBIA = CIUDADES_COLOMBIA_COMPLETA | DEPARTAMENTOS_COLOMBIA

# Prefijos de celular colombiano
PREFIJOS_CELULAR_COLOMBIA: Set[str] = {
    '300', '301', '302', '303', '304', '305',
    '310', '311', '312', '313', '314', '315', '316', '317', '318',
    '320', '321', '322', '323', '324',
    '350', '351',
    '330', '331', '332', '333'
}

# Hojas a excluir silenciosamente
HOJAS_EXCLUIR_SILENCIOSAMENTE: Set[str] = {
    'INSTRUCCIONES', 'INFO', 'DATOS', 'CONTENIDO', 'INDICE', 'ÍNDICE',
    'GUIA DE USO', 'GUÍA DE USO', 'CONTROL DE CAMBIOS', 'HOJA1', 'SHEET1',
    'INSTRUCTIVO', 'PARAMETROS', 'PARÁMETROS', 'CONFIGURACION', 'CONFIGURACIÓN',
    'LISTA', 'LISTAS', 'VALIDACION', 'VALIDACIÓN', 'CATALOGO', 'CATÁLOGO',
    'RESUMEN', 'PORTADA', 'CARATULA', 'CARÁTULA', 'INICIO', 'HOME',
    'MENU', 'MENÚ', 'ANEXO TECNICO', 'ANEXO TÉCNICO', 'GLOSARIO',
}

# Hojas que no generan alerta individualmente pero se mencionan si no hay hoja de servicios
HOJAS_SIN_SERVICIOS_VALIDOS: Set[str] = {
    'PAQUETES', 'TARIFAS PAQUETES', 'PAQUETE',
    'COSTO VIAJE', 'COSTO DE VIAJE', 'COSTOS VIAJE',
}

# Palabras inválidas para CUPS
PALABRAS_INVALIDAS_CUPS: List[str] = [
    'CODIGO', 'CUPS', 'ITEM', 'DESCRIPCION', 'TARIFA', 'TOTAL', 'SUBTOTAL',
    'DEPARTAMENTO', 'MUNICIPIO', 'HABILITACION', 'HABIITACION', 'DIRECCION',
    'TELEFONO', 'EMAIL', 'SEDE', 'NOMBRE', 'NUMERO', 'ESPECIALIDAD',
    'MANUAL', 'OBSERV', 'PORCENTAJE', 'HOMOLOGO', 'N°', 'NO.',
    'NOTA', 'NOTAS', 'ACLARATORIA', 'ACLARATORIAS', 'ACLARACION', 'ACLARACIONES',
    'INCLUYE', 'NO INCLUYE', 'EXCLUYE',
    'USO DE EQUIPO', 'DERECHO DE SALA', 'DERECHO SALA',
    'VER NOTA', 'VER NOTAS', 'SEGUN NOTA',
    'APLICA', 'NO APLICA', 'SEGÚN', 'SEGUN',
    'CONSULTAR', 'REVISAR', 'PENDIENTE',
    'VALOR', 'PRECIO', 'COSTO',
    'CONTRATO', 'ACTA', 'OTROSI', 'OTROSÍ',
    'VIGENTE', 'VIGENCIA',
    'TRASLADO', 'ORIGEN', 'DESTINO',
    'TARIFAS PROPIAS', 'TARIFA PROPIA',
]

# Patrones de direcciones
PATRONES_DIRECCION: List[str] = [
    'CARRERA ', 'CRA ', 'CRA. ', 'CR ', 'CALLE ', 'CL ', 'CL. ',
    'AVENIDA ', 'AV ', 'AV. ', 'DIAGONAL ', 'DG ', 'DG. ',
    'TRANSVERSAL ', 'TV ', 'TV. ', 'KM ', 'KILOMETRO', 'KILÓMETRO',
    'LOCAL ', 'PISO ', 'OFICINA ', 'OF ', 'CONSULTORIO', 'TORRE ',
    'BLOQUE ', 'MANZANA', 'CASA ', 'APARTAMENTO', 'APTO', 'EDIFICIO',
    'CENTRO COMERCIAL', 'C.C.', 'BARRIO ', 'VEREDA ', 'SECTOR '
]

# Patrones inválidos para CUPS
PATRONES_INVALIDOS_CUPS: List[str] = [
    r'^\*',
    r'^-+$',
    r'^\d{1,2}$',
    r'^N\.?A\.?$',
    r'^N/A$',
    r'INCLUYE',
    r'NOTA\s*\d*',
]

# Palabras clave de ambulancias
PALABRAS_AMBULANCIA: Set[str] = {
    'AMBULANCIA', 'TAM', 'TAB', 'TRASLADO ASISTENCIAL',
    'TRANSPORTE ASISTENCIAL', 'SERVICIO AMBULANCIA'
}


# ══════════════════════════════════════════════════════════════════════════════
# FUNCIONES DE DETECCIÓN DE ARCHIVOS v15.1
# ══════════════════════════════════════════════════════════════════════════════

def es_archivo_tarifas_valido(nombre: str) -> Tuple[bool, str]:
    """
    v15.1: Detecta si un archivo es válido para procesamiento de tarifas.
    
    Returns: (es_valido: bool, tipo: str)
    Tipos: 'ANEXO_1', 'TARIFAS', 'OTROSI', 'INVALIDO'
    """
    if not nombre:
        return False, 'INVALIDO'
    
    nombre_upper = nombre.upper()
    
    # EXCLUSIONES: archivos que NO se deben procesar
    palabras_excluir = [
        'MEDICAMENT', 'MEDICAMENTO', 'MEDICAMENTOS',
        'FARMACO', 'FÁRMACO', 'FARMACOS', 'FÁRMACOS',
        'INSUMO', 'INSUMOS'
    ]
    
    for palabra in palabras_excluir:
        if palabra in nombre_upper:
            if 'SERVICIO' in nombre_upper or 'SERV' in nombre_upper:
                continue
            return False, 'INVALIDO'
    
    # v15.1: EXCLUSIÓN: "ANALISIS DE TARIFAS" y variantes NO se procesan
    if re.search(r'AN[AÁ]LISIS\s*(DE\s*)?(TARIFAS?|TARIFA)', nombre_upper):
        return False, 'INVALIDO'
    
    # DETECCIÓN 1: Archivos OTROSÍ (tienen prioridad)
    patrones_otrosi = [
        r'OTRO\s*S[IÍ]\s*[_#\-\s]*(\d+)',
        r'OTROS[IÍ]\s*[_#\-\s]*(\d+)',
        r'OT[_\-\s]?(\d+)',
        r'ADICI[OÓ]N\s*[_#\-\s]*(\d+)',
        r'MODIFICACI[OÓ]N\s*[_#\-\s]*(\d+)',
    ]
    
    for patron in patrones_otrosi:
        if re.search(patron, nombre_upper):
            if 'TARIFA' in nombre_upper or 'ANEXO' in nombre_upper:
                return True, 'OTROSI'
    
    # DETECCIÓN 2: ANEXO 1 (formato tradicional)
    patrones_anexo1 = [
        r'ANEXO\s*[_\-]?\s*1(?!\d)',
        r'ANEXO\s*[_\-]?\s*01(?!\d)',
        r'ANEXO[_\-]1\b',
    ]
    
    for patron in patrones_anexo1:
        if re.search(patron, nombre_upper):
            return True, 'ANEXO_1'
    
    # DETECCIÓN 3: Archivos de TARIFAS v15.0
    patrones_tarifas = [
        r'^\d{3,4}[_\-]TARIFAS?[_\-]',
        r'^TARIFAS?[_\-]',
        r'[_\-]TARIFAS?[_\-]\d{3}',
    ]
    
    for patron in patrones_tarifas:
        if re.search(patron, nombre_upper):
            return True, 'TARIFAS'
    
    return False, 'INVALIDO'


def contiene_anexo1(nombre: str) -> bool:
    """v15.0: Detecta si el nombre corresponde a un archivo procesable de tarifas."""
    es_valido, tipo = es_archivo_tarifas_valido(nombre)
    return es_valido


def extraer_numero_otrosi(nombre: str) -> Optional[int]:
    """v15.0: Extrae el número de otrosí del nombre del archivo."""
    if not nombre:
        return None
    
    nombre_upper = nombre.upper()
    
    patrones = [
        r'OTRO\s*S[IÍ]\s*[_#\-\s]*N?[OÚº°]?\.?\s*(\d+)',
        r'OTROS[IÍ]\s*[_#\-\s]*(\d+)',
        r'OTRO[\s_\-]?SI[\s_\-#]*(\d+)',
        r'OT\s*[_\-\s]?\s*(\d+)',
        r'ADICI[OÓ]N\s*[_#\-\s]*N?[OÚº°]?\.?\s*(\d+)',
        r'MODIFICA(?:CI[OÓ]N)?\s*[_#\-\s]*(\d+)',
    ]
    
    for patron in patrones:
        match = re.search(patron, nombre_upper)
        if match:
            try:
                return int(match.group(1))
            except (ValueError, IndexError):
                continue
    
    return None


def clasificar_tipo_archivo(nombre: str) -> Dict:
    """v15.0: Clasifica un archivo y retorna información completa."""
    resultado = {
        'es_valido': False,
        'tipo': 'INVALIDO',
        'numero_otrosi': None,
        'es_otrosi': False,
        'motivo_exclusion': None
    }
    
    if not nombre:
        resultado['motivo_exclusion'] = 'Nombre vacío'
        return resultado
    
    nombre_upper = nombre.upper()
    
    palabras_excluir = ['MEDICAMENT', 'FARMACO', 'FÁRMACO', 'INSUMO']
    for palabra in palabras_excluir:
        if palabra in nombre_upper:
            if 'SERVICIO' not in nombre_upper and 'SERV' not in nombre_upper:
                resultado['motivo_exclusion'] = f'Archivo de {palabra.lower()}'
                return resultado
    
    num_otrosi = extraer_numero_otrosi(nombre)
    if num_otrosi:
        resultado['numero_otrosi'] = num_otrosi
        resultado['es_otrosi'] = True
    
    es_valido, tipo = es_archivo_tarifas_valido(nombre)
    resultado['es_valido'] = es_valido
    resultado['tipo'] = tipo
    
    if not es_valido:
        resultado['motivo_exclusion'] = 'No coincide con patrones de tarifas'
    
    return resultado


# ══════════════════════════════════════════════════════════════════════════════
# FUNCIONES DE VALIDACIÓN v15.1
# ══════════════════════════════════════════════════════════════════════════════

def es_telefono_celular_colombiano(valor: str) -> bool:
    """v14.1: Detecta si un valor es un teléfono celular colombiano."""
    if not valor:
        return False
    
    valor_str = str(valor).strip()
    if valor_str.endswith('.0'):
        valor_str = valor_str[:-2]
    
    valor_clean = re.sub(r'[^\d]', '', valor_str)
    
    if len(valor_clean) != 10:
        return False
    
    prefijo = valor_clean[:3]
    return prefijo in PREFIJOS_CELULAR_COLOMBIA


def es_telefono_celular(valor: str) -> bool:
    """Alias para compatibilidad."""
    return es_telefono_celular_colombiano(valor)


def es_municipio_o_departamento(texto: str) -> bool:
    """Verifica si el texto es un municipio o departamento."""
    if not texto:
        return False
    texto_upper = normalizar_texto(texto)
    return texto_upper in MUNICIPIOS_COLOMBIA


def es_direccion(texto: str) -> bool:
    """Verifica si el texto parece ser una dirección."""
    if not texto:
        return False
    texto_upper = normalizar_texto(texto)
    return any(p in texto_upper for p in PATRONES_DIRECCION)


def es_numero_sede(valor: str) -> bool:
    """Verifica si parece un número de sede."""
    if not valor:
        return False
    valor_str = str(valor).strip()
    if valor_str.endswith('.0'):
        valor_str = valor_str[:-2]
    if valor_str.isdigit() and len(valor_str) <= 2:
        return True
    return False


def es_dato_de_sede(fila: list) -> bool:
    """Detecta si una fila contiene datos de sede (no servicios)."""
    if not fila or len(fila) < 3:
        return False
    
    fila_texto = ' '.join([str(x).upper() for x in fila[:5] if x is not None])
    
    indicadores_sede = [
        'CODIGO HABILITACION', 'CÓDIGO HABILITACIÓN',
        'NO. SEDE', 'NUMERO SEDE', 'NÚMERO SEDE',
        'DEPARTAMENTO', 'MUNICIPIO', 'DIRECCION', 'DIRECCIÓN',
    ]
    
    contador = sum(1 for ind in indicadores_sede if ind in fila_texto)
    tiene_cups = 'CUPS' in fila_texto or 'CODIGO' in fila_texto
    tiene_ubicacion = any(x in fila_texto for x in ['DEPARTAMENTO', 'MUNICIPIO', 'CIUDAD'])
    
    if contador >= 2:
        return True
    if tiene_ubicacion and not tiene_cups:
        return True
    
    return False


def es_fila_de_traslados(fila: list) -> bool:
    """v14.1: Detecta si una fila contiene información de traslados."""
    if not fila or len(fila) < 3:
        return False
    
    for i, celda in enumerate(fila[:4]):
        if celda:
            celda_str = str(celda).strip()
            if celda_str.endswith('.0'):
                celda_str = celda_str[:-2]
            celda_upper = celda_str.upper()
            if celda_upper in CIUDADES_COLOMBIA_COMPLETA:
                return True
    
    return False


def es_encabezado_seccion_traslados(fila: list) -> bool:
    """v14.1: Detecta si una fila es el encabezado de una sección de traslados."""
    if not fila:
        return False
    
    fila_texto = ' '.join([str(x).upper().strip() for x in fila if x is not None])
    
    indicadores_traslados = [
        'ORIGEN', 'DESTINO', 'MUNICIPIO ORIGEN', 'MUNICIPIO DESTINO',
        'DEPARTAMENTO DESTINO', 'TIPO DE TRASLADO',
    ]
    
    contador = sum(1 for ind in indicadores_traslados if ind in fila_texto)
    tiene_cups = 'CUPS' in fila_texto
    return contador >= 2 and not tiene_cups


def validar_cups(cups: str, fila: list = None) -> bool:
    """v14.1: Validación de CUPS ULTRA estricta."""
    if not cups:
        return False
    
    cups_str = str(cups).strip()
    if cups_str.endswith('.0'):
        cups_str = cups_str[:-2]
    
    cups_u = cups_str.upper()
    
    if not cups_str or len(cups_str) > 15:
        return False
    
    if cups_u in CIUDADES_COLOMBIA_COMPLETA:
        return False
    
    if cups_u in DEPARTAMENTOS_COLOMBIA:
        return False
    
    for palabra in PALABRAS_INVALIDAS_CUPS:
        if palabra in cups_u:
            return False
    
    for patron in PATRONES_INVALIDOS_CUPS:
        if re.search(patron, cups_u):
            return False
    
    cups_digits = re.sub(r'[^\d]', '', cups_str)
    
    if cups_digits and len(cups_digits) >= 7:
        return False
    
    if es_telefono_celular(cups_str):
        return False
    
    if cups_digits and cups_digits == cups_str and 8 <= len(cups_digits) <= 12:
        return False
    
    if cups_u in ['N.A', 'NA', 'N/A', 'N.A.', '-', '--', '---', 'NINGUNO', 'NINGUNA', 'NULL', 'NONE', '']:
        return False
    
    if cups_digits and cups_digits == cups_str:
        if len(cups_digits) < 4:
            return False
    
    if fila and es_fila_de_traslados(fila):
        return False
    
    return True


def validar_tarifa(tarifa, fila: list = None) -> bool:
    """v14.1: Validación mejorada de tarifas."""
    if tarifa is None:
        return True
    
    valor_str = str(tarifa).strip()
    if valor_str.endswith('.0'):
        valor_str = valor_str[:-2]
    
    if es_telefono_celular(valor_str):
        return False
    
    valor_clean = re.sub(r'[^\d]', '', valor_str)
    if valor_clean and 8 <= len(valor_clean) <= 12:
        if fila:
            fila_texto = ' '.join([str(x).upper() for x in fila[:5] if x])
            for depto in DEPARTAMENTOS_COLOMBIA:
                if depto in fila_texto:
                    return False
    
    return True


# ══════════════════════════════════════════════════════════════════════════════
# FUNCIONES UTILITARIAS
# ══════════════════════════════════════════════════════════════════════════════

def normalizar_texto(texto: Any) -> str:
    """Normaliza texto para comparaciones."""
    if texto is None:
        return ''
    s = str(texto).upper().strip()
    for k, v in {'Á': 'A', 'É': 'E', 'Í': 'I', 'Ó': 'O', 'Ú': 'U', 'Ñ': 'N', 'Ü': 'U'}.items():
        s = s.replace(k, v)
    s = re.sub(r'\s+', ' ', s)
    return s


def similitud_texto(a: str, b: str) -> float:
    """Calcula similitud entre dos textos (0.0 a 1.0)."""
    return SequenceMatcher(None, a.upper(), b.upper()).ratio()


def detectar_formato_real(filepath: str) -> str:
    """Detecta el formato REAL de un archivo Excel."""
    try:
        with open(filepath, 'rb') as f:
            header = f.read(8)
        
        if header[:4] == b'PK\x03\x04':
            try:
                with zipfile.ZipFile(filepath, 'r') as z:
                    names = z.namelist()
                    if any('workbook.bin' in n.lower() for n in names):
                        return 'xlsb'
                    elif any('.xml' in n.lower() for n in names):
                        return 'xlsx'
                return 'xlsx'
            except zipfile.BadZipFile:
                return 'zip_corrupt'
        
        if header[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
            return 'xls_old'
        
        return 'unknown'
    except Exception:
        return 'error'


def leer_excel(ruta: str, sheet_name=0, header=0, engine=None):
    """Lee archivo Excel con manejo automático de motor."""
    if engine:
        return pd.read_excel(ruta, engine=engine, sheet_name=sheet_name, header=header)
    
    formato = detectar_formato_real(ruta)
    
    try:
        if formato == 'xlsb':
            return pd.read_excel(ruta, engine='pyxlsb', sheet_name=sheet_name, header=header)
        elif formato == 'xlsx':
            return pd.read_excel(ruta, engine='openpyxl', sheet_name=sheet_name, header=header)
        elif formato == 'xls_old':
            return pd.read_excel(ruta, engine='xlrd', sheet_name=sheet_name, header=header)
    except:
        pass
    
    for eng in ['openpyxl', 'pyxlsb', 'xlrd']:
        try:
            return pd.read_excel(ruta, engine=eng, sheet_name=sheet_name, header=header)
        except:
            continue
    
    raise Exception(f"No se pudo leer: {ruta}")


def obtener_hojas(ruta: str) -> List[str]:
    """Obtiene lista de hojas de un archivo Excel."""
    formato = detectar_formato_real(ruta)
    ext = os.path.splitext(ruta)[1].lower()
    
    if formato == 'xlsb':
        try:
            from pyxlsb import open_workbook
            with open_workbook(ruta) as wb:
                return list(wb.sheets)
        except:
            pass
    elif formato == 'xlsx':
        try:
            from openpyxl import load_workbook
            wb = load_workbook(ruta, read_only=True, data_only=True)
            hojas = wb.sheetnames
            wb.close()
            return hojas
        except:
            pass
    elif formato == 'xls_old':
        try:
            import xlrd
            wb = xlrd.open_workbook(ruta, on_demand=True)
            return wb.sheet_names()
        except:
            pass
    
    try:
        xl = pd.ExcelFile(ruta)
        return xl.sheet_names
    except:
        return []


def leer_hoja_raw(ruta: str, hoja: str, max_filas: int = 20000) -> List[List]:
    """Lee una hoja de Excel como lista de listas."""
    formato = detectar_formato_real(ruta)
    ext = os.path.splitext(ruta)[1].lower()
    
    try:
        if formato == 'xlsb' or ext == '.xlsb':
            from pyxlsb import open_workbook
            datos = []
            with open_workbook(ruta) as wb:
                with wb.get_sheet(hoja) as sheet:
                    for i, row in enumerate(sheet.rows()):
                        if i >= max_filas:
                            break
                        datos.append([cell.v for cell in row])
            return datos
        elif formato == 'xls_old' or ext == '.xls':
            import xlrd
            wb = xlrd.open_workbook(ruta)
            sheet = wb.sheet_by_name(hoja)
            return [[sheet.cell_value(r, c) for c in range(sheet.ncols)]
                    for r in range(min(sheet.nrows, max_filas))]
        else:
            from openpyxl import load_workbook
            wb = load_workbook(ruta, read_only=True, data_only=True)
            sheet = wb[hoja]
            datos = []
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i >= max_filas:
                    break
                datos.append(list(row))
            wb.close()
            return datos
    except Exception:
        try:
            df = leer_excel(ruta, sheet_name=hoja, header=None)
            df = df.head(max_filas)
            return df.values.tolist()
        except:
            return []


def limpiar_codigo(valor) -> Optional[str]:
    """Limpia código eliminando decimales y espacios."""
    if valor is None:
        return None
    texto = str(valor).strip()
    if texto.endswith('.0'):
        texto = texto[:-2]
    return None if not texto or texto.lower() in ('none', 'nan', '') else texto


def limpiar_tarifa(valor) -> Optional[float]:
    """Convierte tarifa a número."""
    if valor is None:
        return None
    try:
        if isinstance(valor, (int, float)):
            return float(valor) if not pd.isna(valor) else None
        texto = str(valor).replace('$', '').replace(',', '').replace(' ', '').strip()
        return float(texto) if texto and texto.lower() not in ('none', 'nan') else None
    except:
        return None


def limpiar_texto(valor) -> Optional[str]:
    """Limpia texto eliminando espacios extras."""
    if valor is None:
        return None
    texto = str(valor).strip()
    return None if not texto or texto.lower() in ('none', 'nan') else texto


def formatear_habilitacion(codigo, sede) -> str:
    """Formatea código de habilitación con sede."""
    if not codigo:
        return "0000000000-01"
    
    c = str(codigo).strip()
    if c.endswith('.0'):
        c = c[:-2]
    
    if re.match(r'^\d{8,12}-\d{1,2}$', c):
        return c
    
    c_limpio = re.sub(r'[^\d]', '', c)
    
    try:
        if sede is None:
            s = 1
        else:
            sede_str = str(sede).strip()
            if sede_str.endswith('.0'):
                sede_str = sede_str[:-2]
            sede_limpia = re.sub(r'[^\d]', '', sede_str)
            if sede_limpia == c_limpio or len(sede_limpia) > 5:
                s = 1
            else:
                s = int(float(sede_str)) if sede_str else 1
    except:
        s = 1
    
    return f"{c_limpio.zfill(10)}-{str(s).zfill(2)}"


# ══════════════════════════════════════════════════════════════════════════════
# BÚSQUEDA INTELIGENTE DE HOJAS v14.1
# ══════════════════════════════════════════════════════════════════════════════

def debe_excluir_hoja_silenciosamente(nombre_hoja: str) -> bool:
    """Verifica si una hoja debe ser excluida SIN generar alerta."""
    if not nombre_hoja:
        return True
    
    nombre_upper = nombre_hoja.upper().strip()
    
    if nombre_upper in HOJAS_EXCLUIR_SILENCIOSAMENTE:
        return True
    
    if nombre_upper in HOJAS_SIN_SERVICIOS_VALIDOS:
        return True
    
    for patron in HOJAS_SIN_SERVICIOS_VALIDOS:
        if patron in nombre_upper:
            return True
    
    return False


def buscar_hoja_servicios_inteligente(hojas: List[str]) -> Tuple[Optional[str], List[Tuple[str, str]]]:
    """v14.1: Busca la hoja de servicios de forma inteligente."""
    if not hojas:
        return None, []
    
    hojas_norm = {h: h.upper().strip() for h in hojas}
    hojas_excluidas_info = []
    
    for hoja, h_norm in hojas_norm.items():
        if h_norm in HOJAS_SIN_SERVICIOS_VALIDOS:
            hojas_excluidas_info.append((hoja, "Hoja de paquetes/costos - No aplica para T25"))
        else:
            for patron in HOJAS_SIN_SERVICIOS_VALIDOS:
                if patron in h_norm:
                    hojas_excluidas_info.append((hoja, "Hoja de paquetes/costos - No aplica para T25"))
                    break
    
    hojas_validas = {h: h_norm for h, h_norm in hojas_norm.items()
                     if not debe_excluir_hoja_silenciosamente(h_norm)}
    
    if not hojas_validas:
        hojas_validas = hojas_norm
    
    # PASO 1: Buscar hoja "SERVICIOS" exacta
    for hoja, h_norm in hojas_validas.items():
        if h_norm.strip() == 'SERVICIOS':
            return hoja, hojas_excluidas_info
    
    # PASO 2: "TARIFAS DE SERVICIOS" sin modificadores
    patrones_exactos = [
        'TARIFAS DE SERVICIOS', 'TARIFA DE SERVICIOS',
        'TARIFAS DE SERV', 'TARIFA DE SERV',
        'TARIFAS DE SERVICIO', 'TARIFA DE SERVICIO',
    ]
    
    for hoja, h_norm in hojas_validas.items():
        h_clean = ' '.join(h_norm.split())
        for patron in patrones_exactos:
            if h_clean == patron or h_clean.startswith(patron):
                if 'COSTO' not in h_clean and 'VIAJE' not in h_clean and 'PAQUETE' not in h_clean:
                    return hoja, hojas_excluidas_info
    
    # PASO 3: TARIFA + SERV (pero no traslados/paquetes)
    for hoja, h_norm in hojas_validas.items():
        if 'TARIFA' in h_norm and 'SERV' in h_norm:
            if 'TRASLADO' not in h_norm and 'PAQUETE' not in h_norm and 'AMBULANCIA' not in h_norm:
                return hoja, hojas_excluidas_info
    
    # PASO 4: SERVICIO (pero no traslados)
    for hoja, h_norm in hojas_validas.items():
        if 'SERVICIO' in h_norm and 'TRASLADO' not in h_norm:
            return hoja, hojas_excluidas_info
    
    # PASO 5: CUPS
    for hoja, h_norm in hojas_validas.items():
        if 'CUPS' in h_norm:
            if not debe_excluir_hoja_silenciosamente(h_norm):
                return hoja, hojas_excluidas_info
    
    # PASO 6: ANEXO 1
    for hoja, h_norm in hojas_validas.items():
        h_clean = h_norm.replace(' ', '').replace('_', '')
        if h_clean in ['ANEXO1', 'ANEXO01']:
            if not debe_excluir_hoja_silenciosamente(h_norm):
                return hoja, hojas_excluidas_info
    
    return None, hojas_excluidas_info


def generar_mensaje_hojas_disponibles(hojas: List[str], hojas_excluidas_info: List[Tuple[str, str]] = None) -> str:
    """v14.1: Genera mensaje con todas las hojas disponibles."""
    if not hojas:
        return "Archivo sin hojas"
    
    hojas_str = ", ".join([f"'{h}'" for h in hojas])
    msg = f"No se encontró hoja de servicios. Hojas disponibles: [{hojas_str}]"
    
    if hojas_excluidas_info:
        paquetes = [h for h, info in hojas_excluidas_info if 'paquete' in info.lower()]
        if paquetes:
            msg += f". NOTA: Se encontraron hojas de PAQUETES ({', '.join(paquetes)}) que NO se procesan en T25."
    
    return msg


def clasificar_hojas(hojas: List[str]) -> Dict[str, List[str]]:
    """Clasifica las hojas disponibles para generar alertas más descriptivas."""
    clasificacion = {
        'servicios': [], 'medicamentos': [], 'traslados': [],
        'ambulancias': [], 'paquetes': [], 'otras': []
    }
    
    for hoja in hojas:
        h = hoja.upper()
        
        if any(x in h for x in ['TARIFA DE SERV', 'TARIFAS DE SERV', 'TARIFAS SERV']):
            if 'MEDICAMENTO' not in h and 'TRASLADO' not in h and 'AMBULANCIA' not in h:
                clasificacion['servicios'].append(hoja)
                continue
        
        es_ambulancia = any(palabra in h for palabra in PALABRAS_AMBULANCIA)
        
        if es_ambulancia:
            clasificacion['ambulancias'].append(hoja)
        elif 'MEDICAMENTO' in h or 'INSUMO' in h:
            clasificacion['medicamentos'].append(hoja)
        elif 'TRASLADO' in h:
            clasificacion['traslados'].append(hoja)
        elif 'PAQUETE' in h:
            clasificacion['paquetes'].append(hoja)
        elif 'SERVICIO' in h and 'MEDICAMENTO' not in h:
            clasificacion['servicios'].append(hoja)
        else:
            clasificacion['otras'].append(hoja)
    
    return clasificacion


def es_archivo_solo_traslados(hojas: List[str]) -> Tuple[bool, str, str]:
    """Verifica si un archivo SOLO contiene hojas de traslados/ambulancias."""
    if not hojas:
        return False, "", ""
    
    cls = clasificar_hojas(hojas)
    
    if cls['servicios']:
        return False, "", ""
    
    tiene_ambulancias = bool(cls['ambulancias'])
    tiene_traslados = bool(cls['traslados'])
    
    if tiene_ambulancias and not tiene_traslados:
        hojas_amb = ', '.join(cls['ambulancias'][:3])
        if len(cls['ambulancias']) > 3:
            hojas_amb += f" (+{len(cls['ambulancias'])-3} más)"
        return True, f"Archivo contiene solo hojas de ambulancias: [{hojas_amb}]", "AMBULANCIAS"
    
    if tiene_traslados and not tiene_ambulancias:
        hojas_traslado = ', '.join(cls['traslados'][:3])
        if len(cls['traslados']) > 3:
            hojas_traslado += f" (+{len(cls['traslados'])-3} más)"
        return True, f"Archivo contiene solo hojas de traslados: [{hojas_traslado}]", "TRASLADOS"
    
    if tiene_ambulancias and tiene_traslados:
        todas = cls['ambulancias'] + cls['traslados']
        hojas_str = ', '.join(todas[:3])
        if len(todas) > 3:
            hojas_str += f" (+{len(todas)-3} más)"
        return True, f"Archivo contiene solo hojas de ambulancias/traslados: [{hojas_str}]", "MIXTO"
    
    return False, "", ""


# ══════════════════════════════════════════════════════════════════════════════
# CLASES DE DATOS
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class Servicio:
    """Representa un servicio extraído del ANEXO 1."""
    codigo_cups: str = ""
    descripcion_cups: str = ""
    codigo_homologo: str = ""
    tarifa_unitaria: Optional[float] = None
    manual_tarifario: str = ""
    porcentaje_tarifario: Optional[float] = None
    codigo_habilitacion: str = ""
    numero_sede: str = "1"
    origen_tarifa: str = ""
    observaciones: str = ""
    archivo_origen: str = ""
    hoja_origen: str = ""
    contrato: str = ""
    ano: str = ""
    
    def to_dict(self) -> Dict:
        return {
            'codigo_cups': self.codigo_cups,
            'descripcion_cups': self.descripcion_cups,
            'codigo_homologo': self.codigo_homologo,
            'tarifa_unitaria': self.tarifa_unitaria,
            'manual_tarifario': self.manual_tarifario,
            'porcentaje_tarifario': self.porcentaje_tarifario,
            'codigo_habilitacion': self.codigo_habilitacion,
            'numero_sede': self.numero_sede,
            'origen_tarifa': self.origen_tarifa,
            'observaciones': self.observaciones,
            'archivo_origen': self.archivo_origen,
            'hoja_origen': self.hoja_origen,
            'contrato': self.contrato,
            'ano': self.ano
        }


@dataclass
class Sede:
    """Representa una sede de prestador."""
    codigo_habilitacion: str = ""
    numero_sede: str = "1"
    nombre: str = ""
    departamento: str = ""
    municipio: str = ""
    direccion: str = ""


@dataclass
class ResultadoProcesamiento:
    """Resultado del procesamiento de un archivo."""
    exito: bool = False
    mensaje: str = ""
    servicios: List[Dict] = field(default_factory=list)
    sedes: List[Dict] = field(default_factory=list)
    alertas: List[Dict] = field(default_factory=list)
    hoja_procesada: str = ""
    total_filas: int = 0
    filas_validas: int = 0


# ══════════════════════════════════════════════════════════════════════════════
# PROCESADOR DE ANEXO
# ══════════════════════════════════════════════════════════════════════════════

class ProcesadorAnexo:
    """Procesador de archivos ANEXO 1 para extracción de servicios. Version 15.1."""
    
    def __init__(self, contrato_id: str = "", ano: str = "", origen_tarifa: str = ""):
        self.contrato_id = contrato_id
        self.ano = ano
        self.origen_tarifa = origen_tarifa
        self.alertas: List[Dict] = []
        self.servicios: List[Servicio] = []
        self.sedes: List[Sede] = []
        self._timeout_flag = False
    
    def agregar_alerta(self, tipo: str, mensaje: str, archivo: str = ""):
        """Agrega una alerta al procesamiento."""
        self.alertas.append({
            'tipo': tipo,
            'mensaje': mensaje,
            'contrato': self.contrato_id,
            'archivo': archivo
        })
    
    def procesar_con_timeout(self, ruta: str, nombre_archivo: str, 
                             timeout: int = 120) -> ResultadoProcesamiento:
        """Procesa archivo con timeout."""
        resultado = ResultadoProcesamiento()
        resultado_holder = {'resultado': resultado}
        
        def worker():
            try:
                resultado_holder['resultado'] = self._procesar_archivo(ruta, nombre_archivo)
            except Exception as e:
                resultado_holder['resultado'].exito = False
                resultado_holder['resultado'].mensaje = f"Error: {str(e)[:100]}"
        
        thread = threading.Thread(target=worker)
        thread.start()
        thread.join(timeout=timeout)
        
        if thread.is_alive():
            self._timeout_flag = True
            self.agregar_alerta('TIMEOUT', f"Timeout procesando archivo ({timeout}s)", nombre_archivo)
            resultado_holder['resultado'].exito = False
            resultado_holder['resultado'].mensaje = f"Timeout ({timeout}s)"
        
        return resultado_holder['resultado']
    
    def _procesar_archivo(self, ruta: str, nombre_archivo: str) -> ResultadoProcesamiento:
        """Procesa un archivo ANEXO 1."""
        resultado = ResultadoProcesamiento()
        
        try:
            hojas = obtener_hojas(ruta)
            if not hojas:
                resultado.mensaje = "No se pudieron obtener hojas del archivo"
                self.agregar_alerta('ERROR_LECTURA', resultado.mensaje, nombre_archivo)
                return resultado
            
            es_solo_traslados, msg_traslados, tipo_traslado = es_archivo_solo_traslados(hojas)
            if es_solo_traslados:
                resultado.mensaje = msg_traslados
                self.agregar_alerta('ARCHIVO_SOLO_TRASLADOS', msg_traslados, nombre_archivo)
                return resultado
            
            hoja_servicios, hojas_excluidas = buscar_hoja_servicios_inteligente(hojas)
            
            if not hoja_servicios:
                msg = generar_mensaje_hojas_disponibles(hojas, hojas_excluidas)
                resultado.mensaje = msg
                self.agregar_alerta('HOJA_NO_ENCONTRADA', msg, nombre_archivo)
                return resultado
            
            resultado.hoja_procesada = hoja_servicios
            
            datos = leer_hoja_raw(ruta, hoja_servicios)
            if not datos:
                resultado.mensaje = f"Hoja '{hoja_servicios}' vacía o no legible"
                self.agregar_alerta('ERROR_LECTURA', resultado.mensaje, nombre_archivo)
                return resultado
            
            resultado.total_filas = len(datos)
            
            columnas = self._detectar_columnas(datos)
            if not columnas:
                resultado.mensaje = "No se detectaron columnas de servicios"
                self.agregar_alerta('COLUMNAS_NO_DETECTADAS', resultado.mensaje, nombre_archivo)
                return resultado
            
            sedes_detectadas = self._detectar_sedes(datos, columnas)
            
            servicios_extraidos = self._extraer_servicios(
                datos, columnas, sedes_detectadas, nombre_archivo, hoja_servicios
            )
            
            resultado.servicios = [s.to_dict() for s in servicios_extraidos]
            resultado.sedes = [{'codigo': s.codigo_habilitacion, 'numero': s.numero_sede} for s in sedes_detectadas]
            resultado.filas_validas = len(servicios_extraidos)
            resultado.alertas = self.alertas
            
            if servicios_extraidos:
                resultado.exito = True
                resultado.mensaje = f"Extraídos {len(servicios_extraidos)} servicios de '{hoja_servicios}'"
            else:
                resultado.mensaje = "No se encontraron servicios válidos"
                self.agregar_alerta('SIN_SERVICIOS', resultado.mensaje, nombre_archivo)
            
            return resultado
            
        except Exception as e:
            resultado.mensaje = f"Error procesando: {str(e)[:100]}"
            self.agregar_alerta('ERROR_PROCESAMIENTO', resultado.mensaje, nombre_archivo)
            return resultado
    
    def _detectar_columnas(self, datos: List[List]) -> Optional[Dict[str, int]]:
        """Detecta las columnas de servicios en los datos."""
        columnas = {}
        
        for i, fila in enumerate(datos[:50]):
            fila_texto = [normalizar_texto(c) for c in fila]
            
            for j, texto in enumerate(fila_texto):
                if 'CUPS' in texto or 'CODIGO' in texto:
                    if 'HABILITACION' not in texto and 'SEDE' not in texto:
                        columnas['cups'] = j
                        columnas['fila_encabezado'] = i
                        break
            
            if 'cups' in columnas:
                for j, texto in enumerate(fila_texto):
                    if 'DESCRIPCION' in texto and 'descripcion' not in columnas:
                        columnas['descripcion'] = j
                    elif 'HOMOLOGO' in texto:
                        columnas['homologo'] = j
                    elif 'TARIFA' in texto and 'tarifa' not in columnas:
                        if 'MANUAL' not in texto and 'PORCENTAJE' not in texto:
                            columnas['tarifa'] = j
                    elif 'MANUAL' in texto:
                        columnas['manual'] = j
                    elif 'PORCENTAJE' in texto or '%' in texto:
                        columnas['porcentaje'] = j
                    elif 'HABILITACION' in texto:
                        columnas['habilitacion'] = j
                    elif 'SEDE' in texto:
                        columnas['sede'] = j
                break
        
        return columnas if 'cups' in columnas else None
    
    def _detectar_sedes(self, datos: List[List], columnas: Dict[str, int]) -> List[Sede]:
        """Detecta las sedes en el archivo."""
        sedes = []
        fila_inicio = columnas.get('fila_encabezado', 0)
        
        for i in range(min(fila_inicio, 30)):
            fila = datos[i]
            for j, celda in enumerate(fila[:6]):
                valor = limpiar_codigo(celda)
                if valor and re.match(r'^\d{10,12}(-\d{1,2})?$', valor):
                    sede = Sede(codigo_habilitacion=valor.split('-')[0])
                    if '-' in valor:
                        sede.numero_sede = valor.split('-')[1]
                    sedes.append(sede)
                    break
        
        if not sedes:
            sedes.append(Sede(codigo_habilitacion="0000000000", numero_sede="1"))
        
        return sedes
    
    def _extraer_servicios(self, datos: List[List], columnas: Dict[str, int],
                          sedes: List[Sede], archivo: str, hoja: str) -> List[Servicio]:
        """Extrae los servicios de los datos."""
        servicios = []
        fila_inicio = columnas.get('fila_encabezado', 0) + 1
        sede_actual = sedes[0] if sedes else Sede()
        en_seccion_traslados = False
        
        for i, fila in enumerate(datos[fila_inicio:], start=fila_inicio):
            if self._timeout_flag:
                break
            
            if es_encabezado_seccion_traslados(fila):
                en_seccion_traslados = True
                continue
            
            if en_seccion_traslados:
                fila_texto = ' '.join([str(c).strip() for c in fila if c])
                if not fila_texto or 'CUPS' in fila_texto.upper():
                    en_seccion_traslados = False
                continue
            
            if es_fila_de_traslados(fila):
                continue
            
            cups_idx = columnas['cups']
            cups_valor = limpiar_codigo(fila[cups_idx]) if cups_idx < len(fila) else None
            
            if not cups_valor:
                continue
            
            if not validar_cups(cups_valor, fila):
                continue
            
            servicio = Servicio(
                codigo_cups=cups_valor,
                contrato=self.contrato_id,
                ano=self.ano,
                origen_tarifa=self.origen_tarifa,
                archivo_origen=archivo,
                hoja_origen=hoja,
                codigo_habilitacion=sede_actual.codigo_habilitacion,
                numero_sede=sede_actual.numero_sede
            )
            
            if 'descripcion' in columnas:
                idx = columnas['descripcion']
                if idx < len(fila):
                    servicio.descripcion_cups = limpiar_texto(fila[idx]) or ""
            
            if 'homologo' in columnas:
                idx = columnas['homologo']
                if idx < len(fila):
                    servicio.codigo_homologo = limpiar_codigo(fila[idx]) or ""
            
            if 'tarifa' in columnas:
                idx = columnas['tarifa']
                if idx < len(fila):
                    tarifa = limpiar_tarifa(fila[idx])
                    if tarifa is not None and validar_tarifa(tarifa, fila):
                        servicio.tarifa_unitaria = tarifa
            
            if 'manual' in columnas:
                idx = columnas['manual']
                if idx < len(fila):
                    servicio.manual_tarifario = limpiar_texto(fila[idx]) or ""
            
            if 'porcentaje' in columnas:
                idx = columnas['porcentaje']
                if idx < len(fila):
                    pct = limpiar_tarifa(fila[idx])
                    if pct is not None:
                        servicio.porcentaje_tarifario = pct
            
            servicios.append(servicio)
        
        return servicios
